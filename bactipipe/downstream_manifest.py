"""Validated manifest contract for post-QC BactiPipe subcohort analyses."""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
from typing import Iterable
import csv
import json
import re


COMMANDS = {"relate", "detect"}
ROLES = {"sample", "reference"}
COHORT_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,79}$")


class DownstreamManifestError(ValueError):
    """Raised when a downstream analysis manifest is incomplete or ambiguous."""


@dataclass(frozen=True)
class DownstreamSample:
    sample_id: str
    display_name: str
    specimen: str = ""
    collection_date: str = ""
    role: str = "sample"
    reference_name: str = ""
    reference_path: str = ""


@dataclass(frozen=True)
class DownstreamCohort:
    cohort_id: str
    command: str
    organism: str
    samples: tuple[DownstreamSample, ...]
    options: dict[str, object]

    @property
    def reference(self) -> DownstreamSample | None:
        return next((sample for sample in self.samples if sample.role == "reference"), None)

    @property
    def signature(self) -> str:
        payload = {
            "schema_version": 1,
            "cohort_id": self.cohort_id,
            "command": self.command,
            "organism": self.organism,
            "options": self.options,
            "samples": [sample.__dict__ for sample in self.samples],
        }
        serialized = json.dumps(payload, sort_keys=True, separators=(",", ":"))
        return sha256(serialized.encode("utf-8")).hexdigest()


_HEADER_ALIASES = {
    "cohort": "cohort_id",
    "cohort id": "cohort_id",
    "cohort_id": "cohort_id",
    "analysis": "command",
    "analysis command": "command",
    "command": "command",
    "organism": "organism",
    "other organism": "other_organism",
    "other_organism": "other_organism",
    "sample": "sample_id",
    "sample id": "sample_id",
    "sample_id": "sample_id",
    "accession number": "sample_id",
    "display name": "display_name",
    "display_name": "display_name",
    "isolate": "display_name",
    "specimen": "specimen",
    "date": "collection_date",
    "collection date": "collection_date",
    "collection_date": "collection_date",
    "role": "role",
    "reference name": "reference_name",
    "reference_name": "reference_name",
    "reference path": "reference_path",
    "reference_path": "reference_path",
    "run ani": "run_ani",
    "run_ani": "run_ani",
    "run ska": "run_ska",
    "run_ska": "run_ska",
    "run cgmlst": "run_cgmlst",
    "run_cgmlst": "run_cgmlst",
    "detect amr": "detect_amr",
    "detect_amr": "detect_amr",
    "detect virulence": "detect_virulence",
    "detect_virulence": "detect_virulence",
    "extended virulence": "extended_virulence",
    "extended_virulence": "extended_virulence",
    "abricate resfinder": "abricate_resfinder",
    "abricate_resfinder": "abricate_resfinder",
    "abricate card": "abricate_card",
    "abricate_card": "abricate_card",
    "minimum identity": "min_identity",
    "min identity": "min_identity",
    "min_identity": "min_identity",
    "minimum coverage": "min_coverage",
    "min coverage": "min_coverage",
    "min_coverage": "min_coverage",
}

_OPTION_TYPES = {
    "run_ani": "bool",
    "run_ska": "bool",
    "run_cgmlst": "bool",
    "detect_amr": "bool",
    "detect_virulence": "bool",
    "extended_virulence": "bool",
    "abricate_resfinder": "bool",
    "abricate_card": "bool",
    "min_identity": "float",
    "min_coverage": "float",
}


def read_downstream_manifest(path: Path | str) -> list[DownstreamCohort]:
    path = Path(path)
    if not path.is_file():
        raise DownstreamManifestError(f"Downstream manifest was not found: {path}")
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        rows = _read_xlsx(path)
    elif path.suffix.casefold() in {".tsv", ".txt"}:
        rows = _read_tsv(path)
    else:
        raise DownstreamManifestError("Downstream manifest must be .tsv, .txt, .xlsx, or .xlsm.")
    return _parse_rows(rows)


def write_command_sample_sheet(cohort: DownstreamCohort, path: Path | str) -> Path:
    """Write the command-specific sample sheet consumed by relate or detect."""

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        if cohort.command == "relate":
            writer.writerow(["sample", "isolate", "specimen", "date"])
        for sample in cohort.samples:
            # ``relate --reference`` inserts the reference with its authoritative
            # path. Keeping it in the ordinary sheet would make an external
            # reference resolve incorrectly under --assemblies-dir.
            if cohort.command == "relate" and sample.role == "reference":
                continue
            writer.writerow(
                [sample.sample_id, sample.display_name, sample.specimen, sample.collection_date]
                if cohort.command == "relate"
                else [sample.sample_id, sample.display_name, sample.specimen]
            )
    return path


def _read_tsv(path: Path) -> list[list[object]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.reader(handle, delimiter="\t"))


def _read_xlsx(path: Path) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency message is environment-specific
        raise DownstreamManifestError("openpyxl is required to read Excel manifests.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    # Guided templates contain Instructions and Examples worksheets. Users may
    # save after viewing one of them, so prefer the contract sheet by name rather
    # than relying on whichever tab happened to be active at save time.
    worksheet = workbook["Manifest"] if "Manifest" in workbook.sheetnames else workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _parse_rows(rows: Iterable[Iterable[object]]) -> list[DownstreamCohort]:
    materialized = [list(row) for row in rows]
    header_index, columns = _find_header(materialized)
    grouped: dict[str, list[dict[str, str]]] = {}
    for row_number, raw in enumerate(materialized[header_index + 1 :], start=header_index + 2):
        values = [_text(value) for value in raw]
        if not any(values):
            continue
        record = {
            name: values[index] if index < len(values) else ""
            for index, name in columns.items()
        }
        if not any(record.get(key) for key in ("cohort_id", "command", "organism", "sample_id")):
            continue
        record["_row_number"] = str(row_number)
        grouped.setdefault(record.get("cohort_id", ""), []).append(record)

    if not grouped:
        raise DownstreamManifestError("The downstream manifest contains no sample rows.")
    return [_build_cohort(cohort_id, records) for cohort_id, records in grouped.items()]


def _find_header(rows: list[list[object]]) -> tuple[int, dict[int, str]]:
    required = {"cohort_id", "command", "organism", "sample_id"}
    for index, row in enumerate(rows[:30]):
        columns = {
            position: canonical
            for position, value in enumerate(row)
            if (canonical := _HEADER_ALIASES.get(_header(value)))
        }
        if required.issubset(columns.values()):
            return index, columns
    raise DownstreamManifestError(
        "Could not find a header containing cohort_id, command, organism, and sample_id."
    )


def _build_cohort(cohort_id: str, records: list[dict[str, str]]) -> DownstreamCohort:
    first_row = records[0].get("_row_number", "?")
    if not cohort_id:
        raise DownstreamManifestError(f"Row {first_row}: cohort_id is required.")
    if not COHORT_ID_PATTERN.fullmatch(cohort_id):
        raise DownstreamManifestError(
            f"Cohort ID {cohort_id!r} may contain only letters, numbers, '.', '_', and '-'."
        )

    command = _single_value(records, "command", cohort_id).casefold()
    organism = _single_value(records, "organism", cohort_id)
    other_organism = _single_value(records, "other_organism", cohort_id)
    if command not in COMMANDS:
        raise DownstreamManifestError(
            f"Cohort {cohort_id!r}: command must be 'relate' or 'detect'."
        )
    if not organism:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: organism is required.")
    if organism.casefold() == "other":
        if not other_organism:
            raise DownstreamManifestError(
                f"Cohort {cohort_id!r}: other_organism is required when organism is Other."
            )
        organism = other_organism
    elif other_organism:
        raise DownstreamManifestError(
            f"Cohort {cohort_id!r}: other_organism may be used only when organism is Other."
        )

    samples = []
    for record in records:
        row_number = record["_row_number"]
        sample_id = record.get("sample_id", "")
        if not sample_id:
            raise DownstreamManifestError(
                f"Row {row_number}: sample_id is required, including for a reference row. "
                "Use a stable reference identifier in that column."
            )
        role = (record.get("role") or "sample").casefold()
        if role not in ROLES:
            raise DownstreamManifestError(f"Row {row_number}: role must be sample or reference.")
        samples.append(
            DownstreamSample(
                sample_id=sample_id,
                display_name=record.get("display_name") or sample_id,
                specimen=record.get("specimen", ""),
                collection_date=record.get("collection_date", ""),
                role=role,
                reference_name=record.get("reference_name", ""),
                reference_path=record.get("reference_path", ""),
            )
        )

    sample_ids = [sample.sample_id for sample in samples]
    if len(sample_ids) != len(set(sample_ids)):
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: duplicate sample IDs are not allowed.")
    references = [sample for sample in samples if sample.role == "reference"]
    if len(references) > 1:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: only one reference is allowed.")
    if command == "detect" and references:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: reference applies only to relate.")
    invalid_references = [
        sample.sample_id
        for sample in samples
        if (sample.reference_name or sample.reference_path) and sample.role != "reference"
    ]
    if invalid_references:
        raise DownstreamManifestError(
            f"Cohort {cohort_id!r}: reference_name and reference_path are allowed only on the reference row."
        )
    conflicting_references = [
        sample.sample_id
        for sample in samples
        if sample.reference_name and sample.reference_path
    ]
    if conflicting_references:
        raise DownstreamManifestError(
            f"Cohort {cohort_id!r}: use reference_name or reference_path, not both."
        )
    if command == "relate" and len(samples) < 2:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: relate requires at least two samples.")

    options = {
        name: _cohort_option(records, name, kind, cohort_id)
        for name, kind in _OPTION_TYPES.items()
    }
    return DownstreamCohort(
        cohort_id=cohort_id,
        command=command,
        organism=organism,
        samples=tuple(samples),
        options={name: value for name, value in options.items() if value is not None},
    )


def _single_value(records: list[dict[str, str]], key: str, cohort_id: str) -> str:
    values = {record.get(key, "").strip() for record in records if record.get(key, "").strip()}
    if len(values) > 1:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: conflicting {key} values.")
    return next(iter(values), "")


def _cohort_option(records, key: str, kind: str, cohort_id: str):
    value = _single_value(records, key, cohort_id)
    if not value:
        return None
    if kind == "bool":
        normalized = value.casefold()
        if normalized in {"yes", "true", "1", "on"}:
            return True
        if normalized in {"no", "false", "0", "off"}:
            return False
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: {key} must be yes or no.")
    try:
        number = float(value)
    except ValueError as exc:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: {key} must be numeric.") from exc
    if not 0 <= number <= 100:
        raise DownstreamManifestError(f"Cohort {cohort_id!r}: {key} must be between 0 and 100.")
    return number


def _header(value: object) -> str:
    return re.sub(r"\s+", " ", _text(value).replace("-", " ")).strip().casefold()


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()
